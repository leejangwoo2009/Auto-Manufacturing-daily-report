import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from tkinter import Tk, Label, Button, StringVar, Radiobutton, Toplevel, Canvas, Frame, Scrollbar, messagebox
try:
    from tkcalendar import Calendar
except ImportError:
    raise ImportError("The 'tkcalendar' 모듈이 설치되어 있지 않습니다. 'pip install tkcalendar'로 설치하세요.")

import re  # 정규식을 사용해 데이터 추출

# 엑셀 파일 생성 함수
def create_excel_file(file_name, shift_type, fct_txt_path, led_txt_path, progress_bar=None):
    """엑셀 테이블 생성 함수"""
    # 새 워크북 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ford Table"

    # 컬럼 너비 조정
    sheet.column_dimensions['A'].width = 20  # '분류' 열
    sheet.column_dimensions['B'].width = 40  # '불량 항목' 열
    for col in range(3, 10):  # '시간대' 및 '합계' 열
        sheet.column_dimensions[chr(64 + col)].width = 25

    # 셀 병합 및 값 설정 함수
    def merge_and_set_value(start_row, start_col, end_row, end_col, value, fill_color=None):
        """셀 병합 및 데이터 삽입"""
        cell = sheet.cell(row=start_row, column=start_col)
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    cell.fill = fill

    # 셀 병합 (헤더 등)
    merge_and_set_value(1, 1, 2, 1, "분류")
    merge_and_set_value(1, 2, 2, 2, "불량 항목")
    merge_and_set_value(1, 3, 1, 8, "시간대")
    merge_and_set_value(1, 9, 2, 9, "합계")
    merge_and_set_value(3, 1, 14, 1, "1차 FCT NG")
    merge_and_set_value(16, 1, 23, 1, "외관 불량")
    merge_and_set_value(24, 1, 27, 1, "2회 불량")
    sheet["A15"].value = "LED"
    sheet["A15"].alignment = Alignment(horizontal="center", vertical="center")

    # 시간대 설정
    if shift_type == "주간":
        shift_hours = [
            "A 시간대 : 08:30:00~10:29:59",
            "B 시간대 : 10:30:00~12:29:59",
            "C 시간대 : 12:30:00~14:29:59",
            "D 시간대 : 14:30:00~16:29:59",
            "E 시간대 : 16:30:00~18:29:59",
            "F 시간대 : 18:30:00~20:29:59"
        ]
        shift_keys = ["A 시간대", "B 시간대", "C 시간대", "D 시간대", "E 시간대", "F 시간대"]
    else:  # 야간
        shift_hours = [
            "A' 시간대 : 20:30:00~22:29:59",
            "B' 시간대 : 22:30:00~00:29:59",
            "C' 시간대 : 00:30:00~02:29:59",
            "D' 시간대 : 02:30:00~04:29:59",
            "E' 시간대 : 04:30:00~06:29:59",
            "F' 시간대 : 06:30:00~08:29:59"
        ]
        shift_keys = ["A' 시간대", "B' 시간대", "C' 시간대", "D' 시간대", "E' 시간대", "F' 시간대"]

    for col_idx, time_label in enumerate(shift_hours, start=3):  # 시간대 데이터 삽입 (C~H)
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = time_label
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 불량 항목 배열
    fault_values = [
        "USB-C 관련 문제", "USB-A 관련 문제", "SW 설치 Or Tester기 parts 문제",
        "제품 Q소자 문제 가능성 많음", "Mini B 관련 문제 or Carplay", "제품 Power Pin 문제",
        "제품 USB-C 문제", "제품 USB-A 문제", "제품 Mini B 문제",
        "회로 문제(암전류)", "테스터기 Power 관련 문제", "제품 충전 프로파일 문제",
        "Vision NG", "외관 불량", "바코드 불량", "실크 불량", "USB-A Pin 불량",
        "USB-C Pin 불량", "Power Pin 불량", "Mini B Pin 불량", "체결부 불량",
        "FCT 2회 NG", "Vision 2회 NG", "외관 2회 NG", "Reflash NG"
    ]

    for row_idx, fault_value in enumerate(fault_values, start=3):
        cell = sheet.cell(row=row_idx, column=2)
        cell.value = fault_value
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 데이터 매트릭스 초기화
    fct_matrix = [[0 for _ in range(6)] for _ in range(24)]
    led_matrix = [[0 for _ in range(6)] for _ in range(24)]

    # FCT NG 데이터 처리
    if os.path.exists(fct_txt_path):
        with open(fct_txt_path, "r", encoding="utf-8") as file:
            for line in file:
                match = re.search(r"(A|B|C|D|E|F)'? 시간대 & FCT\d & ([^:]+) : (\d+)개", line)
                if match:
                    time_slot, fault, count = match.groups()
                    time_slot_key = f"{time_slot}' 시간대" if shift_type == "야간" else f"{time_slot} 시간대"
                    if time_slot_key in shift_keys and fault in fault_values:
                        time_idx = shift_keys.index(time_slot_key)
                        fault_idx = fault_values.index(fault)
                        fct_matrix[fault_idx][time_idx] += int(count)

    # LED NG 데이터 처리
    if os.path.exists(led_txt_path):
        with open(led_txt_path, "r", encoding="utf-8") as file:
            start_reading = False
            for line in file:
                if "======== 시간대별 & LED별 조건별 요약 ========" in line:
                    start_reading = True
                    continue
                if not start_reading:
                    continue
                match = re.search(r"(A|B|C|D|E|F)'? 시간대 & ([^:]+) : (\d+)개", line)
                if match:
                    time_slot, fault, count = match.groups()
                    time_slot_key = f"{time_slot}' 시간대" if shift_type == "야간" else f"{time_slot} 시간대"
                    if time_slot_key in shift_keys and fault in fault_values:
                        time_idx = shift_keys.index(time_slot_key)
                        fault_idx = fault_values.index(fault)
                        led_matrix[fault_idx][time_idx] += int(count)

    # 데이터 총합 및 엑셀 데이터 입력
    for row in range(3, 27):
        for col in range(3, 9):
            total_value = fct_matrix[row - 3][col - 3] + led_matrix[row - 3][col - 3]
            sheet.cell(row=row, column=col).value = "" if total_value == 0 else total_value  # 0이면 빈칸으로 표시
        total_sum = sum(fct_matrix[row - 3]) + sum(led_matrix[row - 3])
        sheet.cell(row=row, column=9).value = "" if total_sum == 0 else total_sum  # 0이면 빈칸으로 표시

    # 엑셀 저장
    workbook.save(file_name)

# 엑셀 데이터를 GUI로 표시하는 함수
def display_excel(file_path):
    """Tkinter로 엑셀 데이터를 GUI로 표시"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    display_window = Toplevel(root)
    display_window.title("생성된 테이블")
    display_window.geometry("777x555")

    canvas = Canvas(display_window)
    frame = Frame(canvas)
    scrollbar = Scrollbar(display_window, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", expand=True, fill="both")
    canvas.create_window((0, 0), window=frame, anchor="nw")

    # 병합된 셀 정보 가져오기
    merged_ranges = sheet.merged_cells.ranges

    # 병합 셀 검사 함수 수정
    def is_merged_cell(row, col):
        for merged_range in merged_ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return merged_range
        return None

    # 테이블 생성
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            # 병합된 셀인지 확인
            merged_range = is_merged_cell(row_idx, col_idx)
            if merged_range and (row_idx, col_idx) != (merged_range.min_row, merged_range.min_col):
                # 병합된 셀 영역 내부의 나머지 칸은 건너뜀
                continue

            # 병합 영역의 크기 설정
            rowspan = (merged_range.max_row - merged_range.min_row + 1) if merged_range else 1
            colspan = (merged_range.max_col - merged_range.min_col + 1) if merged_range else 1

            value = "" if cell_value is None else str(cell_value)
            label = Label(
                frame,
                text=value,
                borderwidth=1,
                relief="solid",
                anchor="center",
                wraplength=100,
                font=("Arial", 8)
            )
            label.grid(row=row_idx - 1, column=col_idx - 1, rowspan=rowspan, columnspan=colspan, sticky="nsew")

    frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

# 분석 및 저장 실행
def save_and_analyze():
    selected_date = cal.get_date()
    shift = shift_type.get()
    fct_txt_file = f"{selected_date.replace('-', '')}_{shift}_FCT NG List.txt"
    led_txt_file = f"{selected_date.replace('-', '')}_{shift}_LED NG List.txt"
    excel_file_name = f"{selected_date.replace('-', '')}_{shift}_TEST NG Table.xlsx"

    fct_txt_path = os.path.join(
        "C:\\Ford A+C Test center_생산 분석 프로그램_Rev02\\Data\\FORD A+C_Data\\FORD A+C FCT NG List", fct_txt_file)
    led_txt_path = os.path.join(
        "C:\\Ford A+C Test center_생산 분석 프로그램_Rev02\\Data\\FORD A+C_Data\\FORD A+C LED NG List", led_txt_file)
    excel_path = os.path.join(
        "C:\\Ford A+C Test center_생산 분석 프로그램_Rev02\\Data\\FORD A+C_Data\\FORD A+C TEST NG Table", excel_file_name)

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    try:
        create_excel_file(excel_path, shift, fct_txt_path, led_txt_path)
        display_excel(excel_path)
    except Exception as e:
        messagebox.showerror("오류", f"오류 발생: {e}")

# GUI 구성
root = Tk()
root.title("Ford A+C 데이터 분석 도구")
Label(root, text="기준 날짜 선택").grid(row=0, column=0, padx=5, pady=5)
cal = Calendar(root, selectmode="day", date_pattern="yyyy-mm-dd")
cal.grid(row=0, column=1, padx=5, pady=5)
Label(root, text="Shift 선택").grid(row=1, column=0, padx=5, pady=5)
shift_type = StringVar(value="주간")
Radiobutton(root, text="주간", variable=shift_type, value="주간").grid(row=1, column=1, sticky="w", padx=5, pady=2)
Radiobutton(root, text="야간", variable=shift_type, value="야간").grid(row=2, column=1, sticky="w", padx=5, pady=2)
Button(root, text="분석 실행", command=save_and_analyze).grid(row=3, column=0, columnspan=2, pady=10)
root.mainloop()