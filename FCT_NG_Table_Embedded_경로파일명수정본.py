
import os
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from tkinter import Label
import re
from collections import OrderedDict

def create_excel_file_for_table(file_name, shift_type, fct_txt_path, led_txt_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ford Table"
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 40
    for col in range(3, 10):
        sheet.column_dimensions[chr(64 + col)].width = 25

    def merge_and_set_value(start_row, start_col, end_row, end_col, value, fill_color=None):
        cell = sheet.cell(row=start_row, column=start_col)
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    cell.fill = fill

    # 셀 병합 (헤더 등)
    merge_and_set_value(1, 1, 2, 1, "분류")
    merge_and_set_value(1, 2, 2, 2, "불량 항목")
    merge_and_set_value(1, 3, 1, 8, "시간대")
    merge_and_set_value(1, 9, 2, 9, "합계")
    merge_and_set_value(3, 1, 14, 1, "1차 FCT NG")
    merge_and_set_value(16, 1, 23, 1, "외관 불량")
    merge_and_set_value(24, 1, 27, 1, "2회 불량")
    sheet["A15"].value = "LED"
    sheet["A15"].alignment = Alignment(horizontal="center", vertical="center")

    if shift_type == "주간":
        shift_hours = [
            "A 시간대 : 08:30:00~10:29:59",
            "B 시간대 : 10:30:00~12:29:59",
            "C 시간대 : 12:30:00~14:29:59",
            "D 시간대 : 14:30:00~16:29:59",
            "E 시간대 : 16:30:00~18:29:59",
            "F 시간대 : 18:30:00~20:29:59"
        ]
        shift_keys = ["A 시간대", "B 시간대", "C 시간대", "D 시간대", "E 시간대", "F 시간대"]
    else:
        shift_hours = [
            "A' 시간대 : 20:30:00~22:29:59",
            "B' 시간대 : 22:30:00~00:29:59",
            "C' 시간대 : 00:30:00~02:29:59",
            "D' 시간대 : 02:30:00~04:29:59",
            "E' 시간대 : 04:30:00~06:29:59",
            "F' 시간대 : 06:30:00~08:29:59"
        ]
        shift_keys = ["A' 시간대", "B' 시간대", "C' 시간대", "D' 시간대", "E' 시간대", "F' 시간대"]

    for col_idx, time_label in enumerate(shift_hours, start=3):
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = time_label
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 불량 항목 배열
    fault_values = [
        "USB-C 관련 문제", "USB-A 관련 문제", "SW 설치 Or Tester기 parts 문제",
        "제품 Q소자 문제 가능성 많음", "Mini B 관련 문제 or Carplay", "제품 Power Pin 문제",
        "제품 USB-C 문제", "제품 USB-A 문제", "제품 Mini B 문제",
        "회로 문제(암전류)", "테스터기 Power 관련 문제", "제품 충전 프로파일 문제",
        "Vision NG", "외관 불량", "바코드 불량", "실크 불량", "USB-A Pin 불량",
        "USB-C Pin 불량", "Power Pin 불량", "Mini B Pin 불량", "체결부 불량",
        "FCT 2회 NG", "Vision 2회 NG", "외관 2회 NG", "Reflash NG"
    ]

    for row_idx, fault_value in enumerate(fault_values, start=3):
        cell = sheet.cell(row=row_idx, column=2)
        cell.value = fault_value
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = len(fault_values)  # 항목 개수(현재 25개)
    fct_matrix = [[0] * 6 for _ in range(rows)]
    led_matrix = [[0] * 6 for _ in range(rows)]

    if os.path.exists(fct_txt_path):
        with open(fct_txt_path, "r", encoding="utf-8") as file:
            for line in file:
                match = re.search(r"(A|B|C|D|E|F)'? 시간대 & FCT\d & ([^:]+) : (\d+)개", line)
                if match:
                    time_slot, fault, count = match.groups()
                    time_slot_key = f"{time_slot}' 시간대" if shift_type == "야간" else f"{time_slot} 시간대"
                    if time_slot_key in shift_keys and fault in fault_values:
                        time_idx = shift_keys.index(time_slot_key)
                        fault_idx = fault_values.index(fault)
                        fct_matrix[fault_idx][time_idx] += int(count)

    if os.path.exists(led_txt_path):
        with open(led_txt_path, "r", encoding="utf-8") as file:
            start_reading = False
            for line in file:
                if "======== 시간대별 & LED별 조건별 요약 ========" in line:
                    start_reading = True
                    continue
                if not start_reading:
                    continue
                match = re.search(r"(A|B|C|D|E|F)'? 시간대 & ([^:]+) : (\d+)개", line)
                if match:
                    time_slot, fault, count = match.groups()
                    time_slot_key = f"{time_slot}' 시간대" if shift_type == "야간" else f"{time_slot} 시간대"
                    if time_slot_key in shift_keys and fault in fault_values:
                        time_idx = shift_keys.index(time_slot_key)
                        fault_idx = fault_values.index(fault)
                        led_matrix[fault_idx][time_idx] += int(count)

    for row in range(3, 3 + len(fault_values)):
        r = row - 3
        for col in range(3, 9):
            total_value = fct_matrix[r][col - 3] + led_matrix[r][col - 3]
            sheet.cell(row=row, column=col).value = "" if total_value == 0 else total_value
        total_sum = sum(fct_matrix[r]) + sum(led_matrix[r])
        sheet.cell(row=row, column=9).value = "" if total_sum == 0 else total_sum

    save_dir = r"C:\Ford A+C Test center_생산 분석 프로그램_Rev02\Data\FORD A+C_Data\FORD A+C TEST NG Table"
    os.makedirs(save_dir, exist_ok=True)
    base_filename = os.path.basename(file_name).replace("table_", "").replace(".xlsx", "")
    new_file_name = os.path.join(save_dir, f"{base_filename}_TEST NG Table.xlsx")
    workbook.save(new_file_name)
    return new_file_name

def display_excel_embedded(file_path, parent_frame):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            merged_range = None
            for r in sheet.merged_cells.ranges:
                if r.min_row <= row_idx <= r.max_row and r.min_col <= col_idx <= r.max_col:
                    merged_range = r
                    break
            if merged_range and (row_idx, col_idx) != (merged_range.min_row, merged_range.min_col):
                continue
            rowspan = (merged_range.max_row - merged_range.min_row + 1) if merged_range else 1
            colspan = (merged_range.max_col - merged_range.min_col + 1) if merged_range else 1

            value = "" if cell_value is None else str(cell_value)
            label = Label(
                parent_frame,
                text=value,
                borderwidth=1,
                relief="solid",
                anchor="center",
                wraplength=100,
                font=("Arial", 8)
            )
            label.grid(row=row_idx - 1, column=col_idx - 1, rowspan=rowspan, columnspan=colspan, sticky="nsew")

